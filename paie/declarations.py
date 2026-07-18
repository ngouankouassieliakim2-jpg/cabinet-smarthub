"""Génération des fichiers de déclaration (CNPS, DISA) — on remplit les fichiers MODÈLES."""
import io
import os
from django.conf import settings
from openpyxl import load_workbook

from .models import BulletinPaie
from .calculs import calculer_bulletin


def _nom_prenoms(employe):
    """Sépare le nom (premier mot) des prénoms (le reste)."""
    parties = (employe.nom_prenoms or "").strip().split()
    if not parties:
        return "", ""
    return parties[0], " ".join(parties[1:])


def generer_cnps_nominative(employeur, mois, annee):
    """Ouvre le fichier MODÈLE CNPS et le remplit avec les salariés du mois (sans le recréer)."""
    chemin_modele = os.path.join(settings.BASE_DIR, "paie", "modeles", "cnps_nominative.xlsx")
    wb = load_workbook(chemin_modele)   # on part du VRAI fichier modèle
    ws = wb.active

    # On écrit les données à partir de la ligne 2 (ligne 1 = en-têtes du modèle, on n'y touche pas)
    bulletins = (BulletinPaie.objects.filter(employe__employeur=employeur, mois=mois, annee=annee)
                 .select_related("employe").order_by("employe__nom_prenoms"))
    ligne = 2
    for b in bulletins:
        e = b.employe
        nom, prenoms = _nom_prenoms(e)
        calcul = calculer_bulletin(b)
        type_sal = "J" if getattr(e, "type_salaire", "mensuel") == "journalier" else "M"
        ws.cell(row=ligne, column=1, value=e.numero_cnps or "")
        ws.cell(row=ligne, column=2, value=nom)
        ws.cell(row=ligne, column=3, value=prenoms)
        ws.cell(row=ligne, column=4, value=e.date_naissance.year if e.date_naissance else "")
        ws.cell(row=ligne, column=5, value=e.date_entree.strftime("%d/%m/%Y") if e.date_entree else "")
        ws.cell(row=ligne, column=6, value=e.date_sortie.strftime("%d/%m/%Y") if e.date_sortie else "NEANT")
        ws.cell(row=ligne, column=7, value=type_sal)
        ws.cell(row=ligne, column=8, value=1)
        ws.cell(row=ligne, column=9, value=round(float(calcul["brut_social"])))
        ws.cell(row=ligne, column=10, value="123")  # branches cotisées : 1=Retraite, 2=Prest. familiales, 3=AT/Maternité
        ligne += 1

    flux = io.BytesIO()
    wb.save(flux)
    flux.seek(0)
    return flux.getvalue()
def generer_disa(employeur, annee):
    """Remplit l'onglet DISA (en-tête + salariés + totaux par formules).
    Gère un nombre variable de salariés en insérant des lignes si besoin."""
    chemin_modele = os.path.join(settings.BASE_DIR, "paie", "modeles", "disa_dasc.xlsm")
    wb = load_workbook(chemin_modele, keep_vba=True)
    ws = wb["DISA"]

    def ecrire(coord, valeur):
        for plage in ws.merged_cells.ranges:
            if coord in plage:
                ws[plage.coord.split(":")[0]] = valeur
                return
        ws[coord] = valeur

    # --- En-tête entreprise (uniquement ces 3 cellules) ---
    ecrire("D8", annee)
    ecrire("D10", employeur.numero_cnps or "0")
    ecrire("H10", (employeur.raison_sociale or "").upper())

    params = getattr(employeur, "parametres", None)
    plafond_annuel = float(getattr(params, "plafond_cnps", 0) or 3375000) * 12

    # --- Salariés de l'année ---
    employe_ids = (BulletinPaie.objects.filter(employe__employeur=employeur, annee=annee)
                   .values_list("employe_id", flat=True).distinct())
    employes = [e for e in employeur.employes.filter(id__in=list(employe_ids)).order_by("nom_prenoms")
                if BulletinPaie.objects.filter(employe=e, annee=annee).exists()]

    PREMIERE_LIGNE = 13
    LIGNES_MODELE = 6            # le modèle prévoit 6 lignes (13 à 18)
    nb = len(employes)

    # Si plus de 6 salariés, on insère les lignes manquantes AVANT les totaux (ligne 19)
    lignes_a_inserer = max(0, nb - LIGNES_MODELE)
    if lignes_a_inserer:
        ws.insert_rows(19, amount=lignes_a_inserer)

    derniere_ligne = PREMIERE_LIGNE + max(nb, LIGNES_MODELE) - 1  # dernière ligne de la zone salariés (avant ajustement)

    # --- Écriture des salariés ---
    ligne = PREMIERE_LIGNE
    ordre = 1
    for e in employes:
        bulletins = list(BulletinPaie.objects.filter(employe=e, annee=annee).order_by("mois"))
        nom, prenoms = _nom_prenoms(e)
        brut_annuel = 0.0
        dernier_brut_mensuel = 0.0
        for b in bulletins:
            c = calculer_bulletin(b)
            brut_annuel += float(c["brut_social"])
            dernier_brut_mensuel = float(c["brut_social"])
        nb_mois = len(bulletins)
        soumis_retraite = min(brut_annuel, plafond_annuel)
        type_sal = "H" if getattr(e, "type_salaire", "mensuel") == "journalier" else "M"

        ws.cell(row=ligne, column=2,  value=ordre)
        ws.cell(row=ligne, column=3,  value=nom or "0")
        ws.cell(row=ligne, column=4,  value=prenoms or "0")
        ws.cell(row=ligne, column=5,  value=e.numero_cnps or "0")
        ws.cell(row=ligne, column=6,  value=e.date_naissance.year if e.date_naissance else 0)
        ws.cell(row=ligne, column=7,  value=e.date_entree.strftime("%d/%m/%Y") if e.date_entree else "0")
        ws.cell(row=ligne, column=8,  value=e.date_sortie.strftime("%d/%m/%Y") if e.date_sortie else "NEANT")
        ws.cell(row=ligne, column=9,  value=type_sal)
        ws.cell(row=ligne, column=10, value=round(dernier_brut_mensuel))
        ws.cell(row=ligne, column=11, value=round(brut_annuel))
        ws.cell(row=ligne, column=12, value=nb_mois)
        ws.cell(row=ligne, column=13, value=round(brut_annuel))
        ws.cell(row=ligne, column=14, value=round(soumis_retraite))
        ws.cell(row=ligne, column=15, value="123")
        ligne += 1
        ordre += 1

    # Si moins de salariés que de lignes modèle, supprimer les lignes vides en trop
    lignes_a_supprimer = max(0, LIGNES_MODELE - nb)
    if lignes_a_supprimer:
        ws.delete_rows(PREMIERE_LIGNE + nb, amount=lignes_a_supprimer)
        derniere_ligne = PREMIERE_LIGNE + nb - 1

    # --- Totaux par FORMULES (on laisse Excel calculer), plages ajustées ---
    ligne_total_page = derniere_ligne + 1          # ligne 19 d'origine, décalée si insertion
    ws.cell(row=ligne_total_page, column=11, value=f"=SUM(K{PREMIERE_LIGNE}:K{derniere_ligne})")   # K
    ws.cell(row=ligne_total_page, column=13, value=f"=SUM(M{PREMIERE_LIGNE}:M{derniere_ligne})")   # M
    ws.cell(row=ligne_total_page, column=14, value=f"=SUM(N{PREMIERE_LIGNE}:N{derniere_ligne})")   # N
    ecrire(f"E{ligne_total_page}", nb)             # effectif de la page

    ligne_total_ent = ligne_total_page + 2         # ligne 21 d'origine (une ligne vide entre les deux)
    ws.cell(row=ligne_total_ent, column=11, value=f"=+K{ligne_total_page}")
    ws.cell(row=ligne_total_ent, column=13, value=f"=+M{ligne_total_page}")
    ws.cell(row=ligne_total_ent, column=14, value=f"=+N{ligne_total_page}")
    ecrire(f"E{ligne_total_ent}", nb)              # effectif de l'entreprise
    # Remplir aussi l'onglet DASC (cotisations calculées par formules + montants saisis)
    if "DASC" in wb.sheetnames:
        remplir_dasc(wb["DASC"], employeur, annee)

    flux = io.BytesIO()
    wb.save(flux)
    flux.seek(0)
    return flux.getvalue()
def remplir_dasc(ws, employeur, annee):
    """Remplit le DASC : en-tête auto + B (déclaré calculé) et E (payé saisi), regroupés par trimestre."""
    from .models import ReglementCNPS
    from datetime import date

    def ecrire(coord, valeur):
        for plage in ws.merged_cells.ranges:
            if coord in plage:
                ws[plage.coord.split(":")[0]] = valeur
                return
        ws[coord] = valeur

    # En-tête automatique
    ecrire("E5", 0)  # CODE ACTIVITE (viendra de l'inscription plus tard)
    ecrire("H5", f" RAISON SOCIALE   :  {(employeur.raison_sociale or '').upper()}")
    ecrire("H6", f"ADRESSE: {getattr(employeur, 'commune', '') or ''}")
    ecrire("H7", f"Matricule C.N.P.S   /    / {employeur.numero_cnps or '0'}")
    ecrire("B8", f"EXERCICE:{annee}")
    commune = getattr(employeur, "commune", "") or "___"
    ecrire("G18", f"     A  {commune.upper()}, le {date.today().strftime('%d/%m/%Y')}")

    # Paiements saisis, par mois -> regroupés en trimestres
    paye_par_mois = {m: 0.0 for m in range(1, 13)}
    for reg in ReglementCNPS.objects.filter(employeur=employeur, annee=annee):
        paye_par_mois[reg.mois] = float(reg.montant_paye or 0)

    # Trimestre -> ligne DASC, et les 3 mois qui le composent
    trimestres = {1: (13, [1, 2, 3]), 2: (16, [4, 5, 6]), 3: (19, [7, 8, 9]), 4: (22, [10, 11, 12])}
    for t, (r, mois_trim) in trimestres.items():
        declare = sum(cotisations_cnps_mois(employeur, annee, m) for m in mois_trim)
        paye = round(sum(paye_par_mois[m] for m in mois_trim))
        ecrire(f"B{r}", round(declare))   # cotisations déclarées du trimestre
        ecrire(f"E{r}", paye)             # paiements du trimestre

def cotisations_cnps_mois(employeur, annee, mois):
    """Total des cotisations CNPS (salariale retraite + patronales) d'un mois."""
    total = 0.0
    bulletins = BulletinPaie.objects.filter(employe__employeur=employeur, annee=annee, mois=mois)
    for b in bulletins:
        c = calculer_bulletin(b)
        total += (float(c["cnps_retraite_salarie"]) + float(c["cnps_retraite_employeur"])
                  + float(c["cnps_prestations_familiales"]) + float(c["cnps_accident_travail"])
                  + float(c["cnps_maternite"]))
    return round(total)